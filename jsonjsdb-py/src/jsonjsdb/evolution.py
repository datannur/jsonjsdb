"""Evolution tracking: compare datasets and log changes."""

from __future__ import annotations

import json
import re
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Literal, cast

import polars as pl
from openpyxl import Workbook

EvolutionType = Literal["add", "delete", "update"]
EVOLUTION_PUBLIC_FIELDS = [
    "timestamp",
    "type",
    "entity",
    "entity_id",
    "parent_entity_id",
    "variable",
    "old_value",
    "new_value",
    "name",
]

VALID_ID_PATTERN = re.compile(r"^[a-zA-Z0-9_, -]+$")
INVALID_ID_PATTERN = re.compile(r"[^a-zA-Z0-9_, -]")


@dataclass
class EvolutionEntry:
    """A single evolution entry representing a change in the database."""

    timestamp: int
    type: EvolutionType
    entity: str
    entity_id: str | int
    parent_entity_id: str | int | None
    parent_entity: str | None
    variable: str | None
    old_value: Any
    new_value: Any
    name: str | None

    def to_dict(self, *, include_internal: bool = False) -> dict[str, Any]:
        """Convert to dict with snake_case keys for JSON output."""
        data = asdict(self)
        if not include_internal:
            data.pop("parent_entity", None)
        return data

    def to_public_row(self) -> list[Any]:
        """Convert to a public evolution row for JSON.js and XLSX output."""
        data = self.to_dict()
        return [data[field] for field in EVOLUTION_PUBLIC_FIELDS]


def _standardize_id(id_value: str) -> str:
    """Standardize ID by removing invalid characters."""
    trimmed = id_value.strip()
    if VALID_ID_PATTERN.match(trimmed):
        return trimmed
    return INVALID_ID_PATTERN.sub("", trimmed)


def _get_parent_info(
    row: dict[str, Any],
    entity: str,
    parent_relations: dict[str, str] | None,
) -> tuple[str | None, str | int | None]:
    """Get parent entity and id based on config or FK convention.

    Args:
        row: Row data
        entity: Current entity/table name
        parent_relations: Mapping of child_table -> parent_table

    Returns:
        Tuple of (parent_entity, parent_entity_id)
    """
    if parent_relations and entity in parent_relations:
        parent_entity = parent_relations[entity]
        fk_col = f"{parent_entity}_id"
        parent_id = row.get(fk_col)
        if isinstance(parent_id, (str, int)):
            return (parent_entity, parent_id)
        return (parent_entity, None)

    # Fallback: auto-detect from first FK column
    for key, value in row.items():
        if key.endswith("_id") and key != "id":
            parent_entity = key[:-3]  # strip "_id"
            if isinstance(value, (str, int)):
                return (parent_entity, value)
        elif key.endswith("Id"):
            parent_entity = key[:-2]  # strip "Id"
            if isinstance(value, (str, int)):
                return (parent_entity, value)
    return (None, None)


def _add_composite_id_if_missing(df: pl.DataFrame) -> tuple[pl.DataFrame, str | None]:
    """Add composite id column if 'id' column is missing.

    Returns tuple of (modified_df, composite name column).
    """
    if df.is_empty():
        return df, None

    if "id" in df.columns:
        return df, None

    columns = df.columns
    if len(columns) < 2:
        raise ValueError("Not enough columns to generate composite id")

    col1, col2 = columns[0], columns[1]
    df_with_id = df.with_columns(
        (pl.col(col1).cast(pl.Utf8) + "---" + pl.col(col2).cast(pl.Utf8)).alias("id")
    )
    return df_with_id, col2


def _df_to_dict_by_id(df: pl.DataFrame) -> dict[str | int, dict[str, Any]]:
    """Convert DataFrame to dict keyed by id column.

    Normalizes id to string for consistent comparison across types.
    """
    result: dict[str | int, dict[str, Any]] = {}
    for row in df.iter_rows(named=True):
        row_id = row.get("id")
        if row_id is not None:
            # Normalize id to string for consistent comparison
            key = str(row_id)
            result[key] = row
    return result


def _values_are_empty(old_val: Any, new_val: Any) -> bool:
    """Check if both values are effectively empty."""
    empty_values = (None, "", [])
    return old_val in empty_values and new_val in empty_values


def _normalize_id_column(df: pl.DataFrame) -> pl.DataFrame:
    """Normalize id column to string type for consistent comparison."""
    if df.is_empty() or "id" not in df.columns:
        return df
    return df.with_columns(pl.col("id").cast(pl.Utf8))


def _get_composite_name(row: dict[str, Any], composite_name_column: str | None) -> Any:
    """Get the display name from the second composite id column."""
    if composite_name_column is None:
        return None
    return row.get(composite_name_column)


def compare_datasets(
    old_df: pl.DataFrame,
    new_df: pl.DataFrame,
    timestamp: int,
    entity: str,
    parent_relations: dict[str, str] | None = None,
) -> list[EvolutionEntry]:
    """Compare two datasets and return list of evolution entries.

    Args:
        old_df: Previous version of the data
        new_df: New version of the data
        timestamp: Unix timestamp in seconds
        entity: Table/entity name
        parent_relations: Mapping of child_table -> parent_table for cascade filtering

    Returns:
        List of EvolutionEntry objects describing the changes
    """
    entries: list[EvolutionEntry] = []

    if entity.startswith("__"):
        return entries

    # Skip tracking for initial creation (no previous data)
    if old_df.is_empty():
        return entries

    # Normalize id columns to string for consistent comparison
    old_df = _normalize_id_column(old_df)
    new_df = _normalize_id_column(new_df)

    old_df, composite_name_column_old = _add_composite_id_if_missing(old_df)
    new_df, composite_name_column_new = _add_composite_id_if_missing(new_df)
    composite_name_column = composite_name_column_new or composite_name_column_old
    has_composite_id = composite_name_column is not None

    map_old = _df_to_dict_by_id(old_df)
    map_new = _df_to_dict_by_id(new_df)

    # Determine all variables to compare
    if new_df.is_empty():
        variables = old_df.columns
    else:
        variables = list(set(old_df.columns) | set(new_df.columns))

    ids_old = set(map_old.keys())
    ids_new = set(map_new.keys())

    ids_added = ids_new - ids_old
    ids_removed = ids_old - ids_new
    common_ids = ids_old & ids_new

    # Detect updates
    for entity_id in common_ids:
        obj_old = map_old[entity_id]
        obj_new = map_new[entity_id]

        for variable in variables:
            if variable == "id":
                continue

            old_value = obj_old.get(variable)
            new_value = obj_new.get(variable)

            if old_value == new_value:
                continue

            if _values_are_empty(old_value, new_value):
                continue

            parent_entity, parent_id = _get_parent_info(
                obj_new, entity, parent_relations
            )
            entries.append(
                EvolutionEntry(
                    timestamp=timestamp,
                    type="update",
                    entity=entity,
                    entity_id=(
                        _standardize_id(str(entity_id))
                        if has_composite_id
                        else entity_id
                    ),
                    parent_entity_id=parent_id,
                    parent_entity=parent_entity,
                    variable=variable,
                    old_value=old_value,
                    new_value=new_value,
                    name=_get_composite_name(obj_new, composite_name_column),
                )
            )

    # Detect additions
    for entity_id in ids_added:
        obj_new = map_new[entity_id]
        parent_entity, parent_id = _get_parent_info(obj_new, entity, parent_relations)
        entries.append(
            EvolutionEntry(
                timestamp=timestamp,
                type="add",
                entity=entity,
                entity_id=(
                    _standardize_id(str(entity_id)) if has_composite_id else entity_id
                ),
                parent_entity_id=parent_id,
                parent_entity=parent_entity,
                variable=None,
                old_value=None,
                new_value=None,
                name=_get_composite_name(obj_new, composite_name_column),
            )
        )

    # Detect deletions
    for entity_id in ids_removed:
        obj_old = map_old[entity_id]
        parent_entity, parent_id = _get_parent_info(obj_old, entity, parent_relations)
        entries.append(
            EvolutionEntry(
                timestamp=timestamp,
                type="delete",
                entity=entity,
                entity_id=(
                    _standardize_id(str(entity_id)) if has_composite_id else entity_id
                ),
                parent_entity_id=parent_id,
                parent_entity=parent_entity,
                variable=None,
                old_value=None,
                new_value=None,
                name=(
                    _get_composite_name(obj_old, composite_name_column)
                    if composite_name_column is not None
                    else obj_old.get("name")
                ),
            )
        )

    return entries


def filter_cascade_entries(entries: list[EvolutionEntry]) -> list[EvolutionEntry]:
    """Filter out cascade add/delete entries where parent has same operation.

    When a parent entity is added or deleted, child entities are also added/deleted.
    This function removes child entries that are part of a cascade operation,
    keeping only the meaningful parent-level changes.

    Args:
        entries: List of evolution entries to filter

    Returns:
        Filtered list with cascade entries removed
    """
    # Index parent operations: (timestamp, type, entity, entity_id)
    parent_ops: set[tuple[int, str, str, str]] = {
        (e.timestamp, e.type, e.entity, str(e.entity_id))
        for e in entries
        if e.type in ("add", "delete")
    }

    result: list[EvolutionEntry] = []
    for entry in entries:
        # Always keep updates
        if entry.type == "update":
            result.append(entry)
            continue

        # Keep entries without parent relation
        if not entry.parent_entity or entry.parent_entity_id is None:
            result.append(entry)
            continue

        # Check if parent has the same operation in this batch
        parent_key = (
            entry.timestamp,
            entry.type,
            entry.parent_entity,
            str(entry.parent_entity_id),
        )
        if parent_key not in parent_ops:
            result.append(entry)

    return result


def load_evolution(path: Path, xlsx_path: Path | None = None) -> list[EvolutionEntry]:
    """Load existing evolution entries.

    If xlsx_path is provided and exists, reads from it (allows user edits).
    Otherwise falls back to evolution.json.
    """
    if xlsx_path and xlsx_path.exists():
        return load_evolution_xlsx(xlsx_path)

    evolution_path = path / "evolution.json"
    if not evolution_path.exists():
        return []

    with open(evolution_path, encoding="utf-8") as f:
        data = json.load(f)

    return [
        EvolutionEntry(
            timestamp=row["timestamp"],
            type=row["type"],
            entity=row["entity"],
            entity_id=row["entity_id"],
            parent_entity_id=row.get("parent_entity_id"),
            parent_entity=row.get("parent_entity"),
            variable=row.get("variable"),
            old_value=row.get("old_value"),
            new_value=row.get("new_value"),
            name=row.get("name"),
        )
        for row in data
    ]


def load_evolution_xlsx(xlsx_path: Path) -> list[EvolutionEntry]:
    """Load evolution entries from an Excel file."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    ws = wb.active
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [str(value) if value is not None else "" for value in rows[0]]
    entries = []
    for row in rows[1:]:  # Skip header
        row_data = dict(zip(headers, row))
        if row_data.get("timestamp") is None:  # Skip empty rows
            continue

        # Parse type with validation
        type_val = str(row_data.get("type")) if row_data.get("type") else "update"
        if type_val not in ("add", "delete", "update"):
            type_val = "update"

        entries.append(
            EvolutionEntry(
                timestamp=int(str(row_data.get("timestamp")))
                if row_data.get("timestamp")
                else 0,
                type=cast(EvolutionType, type_val),
                entity=str(row_data.get("entity")) if row_data.get("entity") else "",
                entity_id=str(row_data.get("entity_id"))
                if row_data.get("entity_id")
                else "",
                parent_entity_id=str(row_data.get("parent_entity_id"))
                if row_data.get("parent_entity_id")
                else None,
                parent_entity=str(row_data.get("parent_entity"))
                if row_data.get("parent_entity")
                else None,
                variable=str(row_data.get("variable"))
                if row_data.get("variable")
                else None,
                old_value=row_data.get("old_value")
                if row_data.get("old_value")
                else None,
                new_value=row_data.get("new_value")
                if row_data.get("new_value")
                else None,
                name=str(row_data.get("name")) if row_data.get("name") else None,
            )
        )
    return entries


def save_evolution(
    entries: list[EvolutionEntry],
    path: Path,
    xlsx_path: Path | None = None,
) -> None:
    """Save evolution entries to JSON and optionally XLSX.

    Args:
        entries: List of evolution entries to save
        path: Directory path for evolution.json
        xlsx_path: Optional path for evolution.xlsx output
    """
    if not entries:
        return

    # Write evolution.json
    evolution_json_path = path / "evolution.json"
    data = [entry.to_dict() for entry in entries]

    with open(evolution_json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, allow_nan=False)
        f.write("\n")

    # Write evolution.json.js
    evolution_jsonjs_path = path / "evolution.json.js"
    rows: list[list[Any]] = [EVOLUTION_PUBLIC_FIELDS]
    for entry in entries:
        rows.append(entry.to_public_row())

    json_array = json.dumps(
        rows, ensure_ascii=False, separators=(",", ":"), allow_nan=False
    )
    content = f"jsonjs.data['evolution'] = {json_array}\n"

    with open(evolution_jsonjs_path, "w", encoding="utf-8") as f:
        f.write(content)

    # Write evolution.xlsx if path provided
    if xlsx_path:
        write_evolution_xlsx(entries, xlsx_path)


def write_evolution_xlsx(entries: list[EvolutionEntry], xlsx_path: Path) -> None:
    """Write evolution entries to an Excel file."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "evolution"

    # Header row
    ws.append(EVOLUTION_PUBLIC_FIELDS)

    # Data rows
    for entry in entries:
        ws.append(
            [
                entry.timestamp,
                entry.type,
                entry.entity,
                str(entry.entity_id) if entry.entity_id is not None else "",
                str(entry.parent_entity_id) if entry.parent_entity_id else "",
                str(entry.variable) if entry.variable else "",
                str(entry.old_value) if entry.old_value is not None else "",
                str(entry.new_value) if entry.new_value is not None else "",
                str(entry.name) if entry.name else "",
            ]
        )

    wb.save(xlsx_path)


def get_timestamp() -> int:
    """Get current Unix timestamp in seconds."""
    return int(time.time())
