"""Tests for evolution tracking."""

import json
import tempfile
from pathlib import Path

import polars as pl
import pytest

from jsonjsdb import Jsonjsdb, Table, EvolutionEntry
from jsonjsdb.evolution import (
    compare_datasets,
    filter_cascade_entries,
    _standardize_id,
    _get_parent_info,
    _df_to_dict_by_id,
    load_evolution,
    load_evolution_xlsx,
    save_evolution,
    get_timestamp,
    write_evolution_xlsx,
)


class TestCompareDatasets:
    """Tests for compare_datasets function."""

    def test_empty_datasets_returns_empty(self):
        """Should return empty list when both datasets are empty."""
        old_df = pl.DataFrame()
        new_df = pl.DataFrame()

        result = compare_datasets(old_df, new_df, 1234567890, "user")
        assert result == []

    def test_skip_internal_entities(self):
        """Should skip entities starting with __."""
        old_df = pl.DataFrame({"id": [1], "name": ["test"]})
        new_df = pl.DataFrame({"id": [2], "name": ["test2"]})

        result = compare_datasets(old_df, new_df, 1234567890, "__table__")
        assert result == []

    def test_detect_addition(self):
        """Should detect added rows."""
        old_df = pl.DataFrame({"id": [1], "name": ["Alice"]})
        new_df = pl.DataFrame({"id": [1, 2], "name": ["Alice", "Bob"]})

        result = compare_datasets(old_df, new_df, 1234567890, "user")

        assert len(result) == 1
        assert result[0].type == "add"
        assert result[0].entity_id == "2"
        assert result[0].entity == "user"

    def test_detect_deletion(self):
        """Should detect deleted rows."""
        old_df = pl.DataFrame({"id": [1, 2], "name": ["Alice", "Bob"]})
        new_df = pl.DataFrame({"id": [1], "name": ["Alice"]})

        result = compare_datasets(old_df, new_df, 1234567890, "user")

        assert len(result) == 1
        assert result[0].type == "delete"
        assert result[0].entity_id == "2"
        assert result[0].name == "Bob"

    def test_detect_all_deletions(self):
        """Should detect when all rows are deleted (new_df empty)."""
        old_df = pl.DataFrame({"id": ["1", "2"], "name": ["Alice", "Bob"]})
        new_df = pl.DataFrame(schema={"id": pl.Utf8, "name": pl.Utf8})

        result = compare_datasets(old_df, new_df, 1234567890, "user")

        assert len(result) == 2
        assert all(e.type == "delete" for e in result)
        assert {e.entity_id for e in result} == {"1", "2"}

    def test_detect_update(self):
        """Should detect updated fields."""
        old_df = pl.DataFrame({"id": [1], "name": ["Alice"], "score": [100]})
        new_df = pl.DataFrame({"id": [1], "name": ["Alice"], "score": [200]})

        result = compare_datasets(old_df, new_df, 1234567890, "user")

        assert len(result) == 1
        assert result[0].type == "update"
        assert result[0].entity_id == "1"
        assert result[0].variable == "score"
        assert result[0].old_value == 100
        assert result[0].new_value == 200

    def test_no_changes(self):
        """Should return empty list when no changes."""
        old_df = pl.DataFrame({"id": [1], "name": ["Alice"]})
        new_df = pl.DataFrame({"id": [1], "name": ["Alice"]})

        result = compare_datasets(old_df, new_df, 1234567890, "user")
        assert result == []

    def test_ignore_empty_value_changes(self):
        """Should ignore changes between null, empty string, and None."""
        old_df = pl.DataFrame({"id": [1], "name": [None]})
        new_df = pl.DataFrame({"id": [1], "name": [""]})

        result = compare_datasets(old_df, new_df, 1234567890, "user")
        assert result == []

    def test_timestamp_format(self):
        """Timestamp should be preserved as provided."""
        old_df = pl.DataFrame({"id": [1], "name": ["Alice"]})
        new_df = pl.DataFrame({"id": [1, 2], "name": ["Alice", "Bob"]})

        timestamp = 1694765432
        result = compare_datasets(old_df, new_df, timestamp, "user")

        assert result[0].timestamp == timestamp
        assert len(str(result[0].timestamp)) == 10  # Unix seconds format

    def test_multiple_changes_same_row(self):
        """Should detect multiple field changes in same row."""
        old_df = pl.DataFrame({"id": [1], "name": ["Alice"], "score": [100]})
        new_df = pl.DataFrame({"id": [1], "name": ["Alicia"], "score": [200]})

        result = compare_datasets(old_df, new_df, 1234567890, "user")

        assert len(result) == 2
        variables = {r.variable for r in result}
        assert variables == {"name", "score"}


class TestCompositeId:
    """Tests for composite ID generation."""

    def test_generate_composite_id(self):
        """Should generate composite id from first two columns when id missing."""
        old_df = pl.DataFrame({"name": ["John"], "email": ["john@test.com"]})
        new_df = pl.DataFrame(
            {
                "name": ["John", "Jane"],
                "email": ["john@test.com", "jane@test.com"],
            }
        )

        result = compare_datasets(old_df, new_df, 1234567890, "user")

        assert len(result) == 1
        assert result[0].type == "add"
        assert "---" in str(result[0].entity_id)
        # No FK column -> parent_entity_id is None
        assert result[0].parent_entity_id is None

    def test_composite_id_with_fk_column(self):
        """Should use FK column for parent_entity_id with composite id."""
        old_df = pl.DataFrame({"variable_id": ["var_1"], "value": ["A"], "freq": [5]})
        new_df = pl.DataFrame(
            {
                "variable_id": ["var_1", "var_2"],
                "value": ["A", "B"],
                "freq": [5, 10],
            }
        )

        result = compare_datasets(
            old_df,
            new_df,
            1234567890,
            "freq",
            parent_relations={"freq": "variable"},
        )

        assert len(result) == 1
        assert result[0].type == "add"
        assert result[0].entity_id == "var_2---B"
        assert result[0].parent_entity_id == "var_2"
        assert result[0].parent_entity == "variable"

    def test_insufficient_columns_raises_error(self):
        """Should raise error when not enough columns for composite id."""
        old_df = pl.DataFrame({"name": ["test"]})
        new_df = pl.DataFrame({"name": ["test"]})

        with pytest.raises(ValueError, match="Not enough columns"):
            compare_datasets(old_df, new_df, 1234567890, "user")


class TestParentEntityId:
    """Tests for parent entity ID detection."""

    def test_detect_parent_id_suffix(self):
        """Should detect parent_id from _id suffix column."""
        old_df = pl.DataFrame(
            {"id": [1, 2], "name": ["A", "B"], "company_id": [10, 20]}
        )
        new_df = pl.DataFrame({"id": [1], "name": ["A"], "company_id": [10]})

        result = compare_datasets(old_df, new_df, 1234567890, "user")

        assert len(result) == 1
        assert result[0].type == "delete"
        assert result[0].parent_entity_id == 20
        assert result[0].parent_entity == "company"

    def test_get_parent_info_no_fk_column(self):
        """Should return None when no FK column exists."""
        row = {"id": 1, "name": "test", "score": 100}
        parent_entity, parent_id = _get_parent_info(row, "user", None)
        assert parent_entity is None
        assert parent_id is None

    def test_get_parent_info_non_str_int_value(self):
        """Should return None for parent_id when FK value is not str or int."""
        row = {"id": 1, "name": "test", "company_id": None}
        parent_entity, parent_id = _get_parent_info(row, "user", None)
        assert parent_entity is None
        assert parent_id is None

        row2 = {"id": 1, "name": "test", "company_id": ["list", "value"]}
        parent_entity2, parent_id2 = _get_parent_info(row2, "user", None)
        assert parent_entity2 is None
        assert parent_id2 is None

    def test_get_parent_info_with_Id_suffix(self):
        """Should detect FK with camelCase Id suffix."""
        row = {"id": 1, "name": "test", "companyId": 42}
        parent_entity, parent_id = _get_parent_info(row, "user", None)
        assert parent_entity == "company"
        assert parent_id == 42

    def test_get_parent_info_with_config(self):
        """Should use config when provided."""
        row = {"id": 1, "name": "test", "dataset_id": 5, "user_id": 10}
        parent_relations = {"variable": "dataset"}
        parent_entity, parent_id = _get_parent_info(row, "variable", parent_relations)
        assert parent_entity == "dataset"
        assert parent_id == 5

    def test_get_parent_info_config_missing_fk_value(self):
        """Should return parent_entity but None parent_id when FK column missing."""
        row = {"id": 1, "name": "test"}
        parent_relations = {"variable": "dataset"}
        parent_entity, parent_id = _get_parent_info(row, "variable", parent_relations)
        assert parent_entity == "dataset"
        assert parent_id is None


class TestStandardizeId:
    """Tests for ID standardization."""

    def test_valid_id_unchanged(self):
        """Valid IDs should remain unchanged."""
        assert _standardize_id("abc123") == "abc123"
        assert _standardize_id("my-id_1") == "my-id_1"

    def test_strip_whitespace(self):
        """Should strip leading/trailing whitespace."""
        assert _standardize_id("  abc  ") == "abc"

    def test_remove_invalid_chars(self):
        """Should remove invalid characters."""
        assert _standardize_id("test@email.com") == "testemailcom"


class TestEvolutionPersistence:
    """Tests for loading and saving evolution data."""

    def test_save_and_load_evolution(self):
        """Should save and load evolution entries correctly."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir)

            entries = [
                EvolutionEntry(
                    timestamp=1234567890,
                    type="add",
                    entity="user",
                    entity_id=1,
                    parent_entity_id=None,
                    parent_entity=None,
                    variable=None,
                    old_value=None,
                    new_value=None,
                    name=None,
                ),
                EvolutionEntry(
                    timestamp=1234567891,
                    type="update",
                    entity="user",
                    entity_id=1,
                    parent_entity_id=None,
                    parent_entity=None,
                    variable="score",
                    old_value=100,
                    new_value=200,
                    name=None,
                ),
            ]

            save_evolution(entries, path)

            # Verify JSON file exists
            assert (path / "evolution.json").exists()
            assert (path / "evolution.json.js").exists()

            # Load and verify
            loaded = load_evolution(path)
            assert len(loaded) == 2
            assert loaded[0].type == "add"
            assert loaded[1].type == "update"
            assert loaded[1].old_value == 100

    def test_load_nonexistent_returns_empty(self):
        """Should return empty list when no evolution file exists."""
        with tempfile.TemporaryDirectory() as tmpdir:
            loaded = load_evolution(Path(tmpdir))
            assert loaded == []


class TestDatabaseEvolutionTracking:
    """Integration tests for evolution tracking in Jsonjsdb."""

    def test_save_tracks_changes_by_default(self):
        """Should track changes by default when saving."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir)

            # Create and save initial data
            db1 = Jsonjsdb()
            db1._tables["user"] = Table("user", db1)
            db1["user"]._df = pl.DataFrame({"id": [1], "name": ["Alice"]})
            db1.save(path, track_evolution=False)  # First save without tracking

            # Modify and save again
            db2 = Jsonjsdb(path)
            db2["user"]._df = pl.DataFrame({"id": [1, 2], "name": ["Alice", "Bob"]})
            db2.save(path)

            # Verify evolution was tracked
            assert (path / "evolution.json").exists()
            with open(path / "evolution.json") as f:
                evolution = json.load(f)
            assert len(evolution) == 1
            assert evolution[0]["type"] == "add"
            assert evolution[0]["entity_id"] == "2"

    def test_save_opt_out_tracking(self):
        """Should not track when track_evolution=False."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir)

            db = Jsonjsdb()
            db._tables["user"] = Table("user", db)
            db["user"]._df = pl.DataFrame({"id": [1], "name": ["Alice"]})
            db.save(path, track_evolution=False)

            # No evolution file should exist
            assert not (path / "evolution.json").exists()

    def test_evolution_accumulates(self):
        """Should append new entries to existing evolution."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir)

            # First save
            db1 = Jsonjsdb()
            db1._tables["user"] = Table("user", db1)
            db1["user"]._df = pl.DataFrame({"id": [1], "name": ["Alice"]})
            db1.save(path, track_evolution=False)

            # Second save with tracking
            db2 = Jsonjsdb(path)
            db2["user"]._df = pl.DataFrame({"id": [1, 2], "name": ["Alice", "Bob"]})
            db2.save(path)

            # Third save with more changes
            db3 = Jsonjsdb(path)
            db3["user"]._df = pl.DataFrame(
                {"id": [1, 2, 3], "name": ["Alice", "Bob", "Charlie"]}
            )
            db3.save(path)

            # Should have 2 entries now
            with open(path / "evolution.json") as f:
                evolution = json.load(f)
            assert len(evolution) == 2

    def test_save_to_different_path_loads_from_disk(self):
        """Should load existing data from disk when saving to different path."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path_a = Path(tmpdir) / "a"
            path_b = Path(tmpdir) / "b"

            # Create data in path_a
            db1 = Jsonjsdb()
            db1._tables["user"] = Table("user", db1)
            db1["user"]._df = pl.DataFrame({"id": ["1"], "name": ["Alice"]})
            db1.save(path_a, track_evolution=False)

            # Create different data in path_b
            db2 = Jsonjsdb()
            db2._tables["user"] = Table("user", db2)
            db2["user"]._df = pl.DataFrame({"id": ["1"], "name": ["Bob"]})
            db2.save(path_b, track_evolution=False)

            # Load from path_a
            db3 = Jsonjsdb(path_a)
            # Modify and save to path_b (should compare with path_b data, not path_a)
            db3["user"]._df = pl.DataFrame({"id": ["1"], "name": ["Charlie"]})
            db3.save(path_b)

            # Evolution should show change from Bob to Charlie (not Alice to Charlie)
            with open(path_b / "evolution.json") as f:
                evolution = json.load(f)
            assert len(evolution) == 1
            assert evolution[0]["type"] == "update"
            assert evolution[0]["old_value"] == "Bob"
            assert evolution[0]["new_value"] == "Charlie"


class TestGetTimestamp:
    """Tests for timestamp helper."""

    def test_returns_unix_seconds(self):
        """Should return current time in Unix seconds."""
        ts = get_timestamp()
        assert isinstance(ts, int)
        assert len(str(ts)) == 10  # Unix seconds have 10 digits


class TestEvolutionXlsx:
    """Tests for Excel export functionality."""

    def test_save_evolution_with_xlsx(self):
        """Should save evolution.xlsx when path provided."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir)
            xlsx_path = path / "evolution.xlsx"

            entries = [
                EvolutionEntry(
                    timestamp=1234567890,
                    type="add",
                    entity="user",
                    entity_id="1",
                    parent_entity_id=None,
                    parent_entity=None,
                    variable=None,
                    old_value=None,
                    new_value=None,
                    name=None,
                ),
            ]

            save_evolution(entries, path, xlsx_path)

            assert xlsx_path.exists()

    def test_write_evolution_xlsx_creates_file(self):
        """Should create xlsx file with correct structure."""
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / "test.xlsx"

            entries = [
                EvolutionEntry(
                    timestamp=1234567890,
                    type="update",
                    entity="user",
                    entity_id="1",
                    parent_entity_id="10",
                    parent_entity="company",
                    variable="score",
                    old_value=100,
                    new_value=200,
                    name="Alice",
                ),
            ]

            write_evolution_xlsx(entries, xlsx_path)

            assert xlsx_path.exists()

            # Verify content using openpyxl
            from openpyxl import load_workbook

            wb = load_workbook(xlsx_path)
            ws = wb.active
            assert ws is not None

            # Check headers
            headers = [cell.value for cell in ws[1]]
            assert headers == [
                "timestamp",
                "type",
                "entity",
                "entity_id",
                "parent_entity_id",
                "parent_entity",
                "variable",
                "old_value",
                "new_value",
                "name",
            ]

            # Check data row
            row = [cell.value for cell in ws[2]]
            assert row[0] == 1234567890
            assert row[1] == "update"
            assert row[2] == "user"
            assert row[3] == "1"
            assert row[4] == "10"
            assert row[5] == "company"
            assert row[6] == "score"
            assert row[7] == "100"
            assert row[8] == "200"
            assert row[9] == "Alice"

    def test_save_evolution_empty_entries_does_nothing(self):
        """Should not create any files when entries list is empty."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir)

            save_evolution([], path)

            assert not (path / "evolution.json").exists()
            assert not (path / "evolution.json.js").exists()

    def test_load_evolution_from_xlsx(self):
        """Should load from xlsx when path provided and exists."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir)
            xlsx_path = path / "evolution.xlsx"

            # Create xlsx with entries
            entries = [
                EvolutionEntry(
                    timestamp=1234567890,
                    type="add",
                    entity="user",
                    entity_id="1",
                    parent_entity_id=None,
                    parent_entity=None,
                    variable=None,
                    old_value=None,
                    new_value=None,
                    name=None,
                ),
            ]
            write_evolution_xlsx(entries, xlsx_path)

            # Also create json with different content
            json_entries = [
                EvolutionEntry(
                    timestamp=9999999999,
                    type="delete",
                    entity="other",
                    entity_id="99",
                    parent_entity_id=None,
                    parent_entity=None,
                    variable=None,
                    old_value=None,
                    new_value=None,
                    name=None,
                ),
            ]
            save_evolution(json_entries, path)

            # Load should prefer xlsx over json
            loaded = load_evolution(path, xlsx_path)
            assert len(loaded) == 1
            assert loaded[0].type == "add"
            assert loaded[0].entity == "user"

    def test_load_evolution_falls_back_to_json(self):
        """Should fall back to json when xlsx not provided."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir)

            entries = [
                EvolutionEntry(
                    timestamp=1234567890,
                    type="update",
                    entity="user",
                    entity_id="1",
                    parent_entity_id=None,
                    parent_entity=None,
                    variable="name",
                    old_value="Old",
                    new_value="New",
                    name=None,
                ),
            ]
            save_evolution(entries, path)

            # Load without xlsx_path
            loaded = load_evolution(path)
            assert len(loaded) == 1
            assert loaded[0].type == "update"

    def test_load_evolution_xlsx_empty_file(self):
        """Should return empty list for xlsx with only headers."""
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / "empty.xlsx"

            # Create xlsx with only headers
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            assert ws is not None
            ws.append(
                [
                    "timestamp",
                    "type",
                    "entity",
                    "entity_id",
                    "parent_entity_id",
                    "parent_entity",
                    "variable",
                    "old_value",
                    "new_value",
                    "name",
                ]
            )
            wb.save(xlsx_path)

            loaded = load_evolution_xlsx(xlsx_path)
            assert loaded == []

    def test_user_can_edit_xlsx_and_changes_are_preserved(self):
        """Integration test: user edits xlsx, changes are preserved on next save."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir)
            xlsx_path = path / "evolution.xlsx"

            # First save creates evolution
            db1 = Jsonjsdb()
            db1._tables["user"] = Table("user", db1)
            db1["user"]._df = pl.DataFrame({"id": ["1"], "name": ["Alice"]})
            db1.save(path, track_evolution=False)

            db2 = Jsonjsdb(path)
            db2["user"]._df = pl.DataFrame({"id": ["1", "2"], "name": ["Alice", "Bob"]})
            db2.save(path, evolution_xlsx=xlsx_path)

            # Verify we have 1 entry
            with open(path / "evolution.json") as f:
                evolution = json.load(f)
            assert len(evolution) == 1

            # User "edits" xlsx - removes the entry (by creating new xlsx with no entries)
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            assert ws is not None
            ws.append(
                [
                    "timestamp",
                    "type",
                    "entity",
                    "entity_id",
                    "parent_entity_id",
                    "parent_entity",
                    "variable",
                    "old_value",
                    "new_value",
                    "name",
                ]
            )
            wb.save(xlsx_path)

            # Third save should read from cleaned xlsx and add new entries
            db3 = Jsonjsdb(path)
            db3["user"]._df = pl.DataFrame(
                {"id": ["1", "2", "3"], "name": ["Alice", "Bob", "Charlie"]}
            )
            db3.save(path, evolution_xlsx=xlsx_path)

            # Should only have 1 entry (the new add), not 2
            with open(path / "evolution.json") as f:
                evolution = json.load(f)
            assert len(evolution) == 1
            assert evolution[0]["entity_id"] == "3"


class TestEdgeCases:
    """Tests for edge cases in evolution tracking."""

    def test_all_rows_deleted(self):
        """Should detect when all rows are deleted."""
        old_df = pl.DataFrame({"id": ["1", "2"], "name": ["Alice", "Bob"]})
        new_df = pl.DataFrame({"id": [], "name": []}).cast(
            {"id": pl.Utf8, "name": pl.Utf8}
        )

        result = compare_datasets(old_df, new_df, 1234567890, "user")

        assert len(result) == 2
        assert all(r.type == "delete" for r in result)
        entity_ids = {r.entity_id for r in result}
        assert entity_ids == {"1", "2"}

    def test_initial_creation_not_tracked(self):
        """Should not track changes when old table is empty (initial creation)."""
        old_df = pl.DataFrame({"id": [], "name": []}).cast(
            {"id": pl.Utf8, "name": pl.Utf8}
        )
        new_df = pl.DataFrame({"id": ["1", "2"], "name": ["Alice", "Bob"]})

        result = compare_datasets(old_df, new_df, 1234567890, "user")

        # No entries generated for initial creation
        assert len(result) == 0

    def test_df_to_dict_by_id_skips_null_ids(self):
        """Should skip rows with null id values."""
        df = pl.DataFrame({"id": ["1", None, "3"], "name": ["A", "B", "C"]})
        result = _df_to_dict_by_id(df)

        assert len(result) == 2
        assert "1" in result
        assert "3" in result


class TestLoadEvolutionXlsxEdgeCases:
    """Edge case tests for load_evolution_xlsx to reach 100% coverage."""

    def test_load_xlsx_with_empty_rows_skips_them(self):
        """Should skip rows where timestamp (first cell) is None."""
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / "with_empty.xlsx"

            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            assert ws is not None
            # Header
            ws.append(
                [
                    "timestamp",
                    "type",
                    "entity",
                    "entity_id",
                    "parent_entity_id",
                    "parent_entity",
                    "variable",
                    "old_value",
                    "new_value",
                    "name",
                ]
            )
            # Valid row
            ws.append(
                [1234567890, "add", "user", "1", None, None, None, None, None, "Alice"]
            )
            # Empty row (timestamp is None)
            ws.append(
                [None, "update", "user", "2", None, None, "name", "old", "new", "Bob"]
            )
            # Another valid row
            ws.append(
                [1234567891, "delete", "user", "3", None, None, None, None, None, "Eve"]
            )

            wb.save(xlsx_path)

            loaded = load_evolution_xlsx(xlsx_path)

            # Should have 2 entries (skipping the empty row)
            assert len(loaded) == 2
            assert loaded[0].entity_id == "1"
            assert loaded[1].entity_id == "3"

    def test_load_xlsx_with_invalid_type_uses_update_fallback(self):
        """Should use 'update' as fallback for invalid type values."""
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / "invalid_type.xlsx"

            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            assert ws is not None
            # Header
            ws.append(
                [
                    "timestamp",
                    "type",
                    "entity",
                    "entity_id",
                    "parent_entity_id",
                    "parent_entity",
                    "variable",
                    "old_value",
                    "new_value",
                    "name",
                ]
            )
            # Row with invalid type
            ws.append(
                [
                    1234567890,
                    "invalid_type",
                    "user",
                    "1",
                    None,
                    None,
                    "name",
                    "old",
                    "new",
                    None,
                ]
            )
            # Row with another invalid type
            ws.append(
                [
                    1234567891,
                    "modify",
                    "user",
                    "2",
                    None,
                    None,
                    "email",
                    "a@b.com",
                    "x@y.com",
                    None,
                ]
            )

            wb.save(xlsx_path)

            loaded = load_evolution_xlsx(xlsx_path)

            assert len(loaded) == 2
            # Both should have type "update" as fallback
            assert loaded[0].type == "update"
            assert loaded[1].type == "update"

    def test_load_xlsx_with_no_active_worksheet(self):
        """Should return empty list when workbook has no active worksheet."""
        from unittest.mock import MagicMock, patch

        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = Path(tmpdir) / "test.xlsx"

            # Create a valid xlsx file first (required for load_workbook to work)
            from openpyxl import Workbook

            wb = Workbook()
            wb.save(xlsx_path)

            # Mock load_workbook to return a workbook with no active sheet
            mock_wb = MagicMock()
            mock_wb.active = None

            with patch("openpyxl.load_workbook", return_value=mock_wb):
                loaded = load_evolution_xlsx(xlsx_path)
                assert loaded == []


class TestFilterCascadeEntries:
    """Tests for cascade filtering of evolution entries."""

    def test_filter_cascade_removes_child_add_when_parent_added(self):
        """Should remove child add entries when parent has same add operation."""
        entries = [
            EvolutionEntry(
                timestamp=1234567890,
                type="add",
                entity="dataset",
                entity_id="ds_1",
                parent_entity_id=None,
                parent_entity=None,
                variable=None,
                old_value=None,
                new_value=None,
                name="Dataset 1",
            ),
            EvolutionEntry(
                timestamp=1234567890,
                type="add",
                entity="variable",
                entity_id="var_1",
                parent_entity_id="ds_1",
                parent_entity="dataset",
                variable=None,
                old_value=None,
                new_value=None,
                name="Variable 1",
            ),
            EvolutionEntry(
                timestamp=1234567890,
                type="add",
                entity="variable",
                entity_id="var_2",
                parent_entity_id="ds_1",
                parent_entity="dataset",
                variable=None,
                old_value=None,
                new_value=None,
                name="Variable 2",
            ),
        ]

        result = filter_cascade_entries(entries)

        # Only the parent add should remain
        assert len(result) == 1
        assert result[0].entity == "dataset"
        assert result[0].entity_id == "ds_1"

    def test_filter_cascade_removes_child_delete_when_parent_deleted(self):
        """Should remove child delete entries when parent is deleted."""
        entries = [
            EvolutionEntry(
                timestamp=1234567890,
                type="delete",
                entity="dataset",
                entity_id="ds_1",
                parent_entity_id=None,
                parent_entity=None,
                variable=None,
                old_value=None,
                new_value=None,
                name="Dataset 1",
            ),
            EvolutionEntry(
                timestamp=1234567890,
                type="delete",
                entity="variable",
                entity_id="var_1",
                parent_entity_id="ds_1",
                parent_entity="dataset",
                variable=None,
                old_value=None,
                new_value=None,
                name="Variable 1",
            ),
        ]

        result = filter_cascade_entries(entries)

        # Only the parent delete should remain
        assert len(result) == 1
        assert result[0].entity == "dataset"

    def test_filter_cascade_keeps_updates(self):
        """Should always keep update entries."""
        entries = [
            EvolutionEntry(
                timestamp=1234567890,
                type="add",
                entity="dataset",
                entity_id="ds_1",
                parent_entity_id=None,
                parent_entity=None,
                variable=None,
                old_value=None,
                new_value=None,
                name=None,
            ),
            EvolutionEntry(
                timestamp=1234567890,
                type="update",
                entity="variable",
                entity_id="var_1",
                parent_entity_id="ds_1",
                parent_entity="dataset",
                variable="name",
                old_value="Old",
                new_value="New",
                name=None,
            ),
        ]

        result = filter_cascade_entries(entries)

        # Both should remain: parent add + child update
        assert len(result) == 2

    def test_filter_cascade_keeps_orphan_entries(self):
        """Should keep entries without parent relation."""
        entries = [
            EvolutionEntry(
                timestamp=1234567890,
                type="add",
                entity="user",
                entity_id="u_1",
                parent_entity_id=None,
                parent_entity=None,
                variable=None,
                old_value=None,
                new_value=None,
                name="User 1",
            ),
        ]

        result = filter_cascade_entries(entries)

        assert len(result) == 1

    def test_filter_cascade_keeps_child_add_when_parent_not_added(self):
        """Should keep child add when parent is not being added."""
        entries = [
            EvolutionEntry(
                timestamp=1234567890,
                type="add",
                entity="variable",
                entity_id="var_1",
                parent_entity_id="ds_existing",
                parent_entity="dataset",
                variable=None,
                old_value=None,
                new_value=None,
                name="New variable on existing dataset",
            ),
        ]

        result = filter_cascade_entries(entries)

        # Should remain because parent dataset is not in the entries
        assert len(result) == 1


class TestParentRelationsConfig:
    """Tests for parent_relations configuration in save()."""

    def test_save_with_parent_relations_filters_cascade(self):
        """Should filter cascade entries when parent_relations provided."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir)

            # Create initial data
            db1 = Jsonjsdb()
            db1._tables["dataset"] = Table("dataset", db1)
            db1._tables["variable"] = Table("variable", db1)
            db1["dataset"]._df = pl.DataFrame({"id": ["ds_1"], "name": ["Dataset 1"]})
            db1["variable"]._df = pl.DataFrame(
                {"id": ["var_1"], "name": ["Var 1"], "dataset_id": ["ds_1"]}
            )
            db1.save(path, track_evolution=False)

            # Add new dataset with variables
            db2 = Jsonjsdb(path)
            db2["dataset"]._df = pl.DataFrame(
                {"id": ["ds_1", "ds_2"], "name": ["Dataset 1", "Dataset 2"]}
            )
            db2["variable"]._df = pl.DataFrame(
                {
                    "id": ["var_1", "var_2", "var_3"],
                    "name": ["Var 1", "Var 2", "Var 3"],
                    "dataset_id": ["ds_1", "ds_2", "ds_2"],
                }
            )
            db2.save(
                path,
                parent_relations={"variable": "dataset"},
            )

            # Load evolution
            with open(path / "evolution.json") as f:
                evolution = json.load(f)

            # Should only have ds_2 add, not var_2 and var_3 adds
            assert len(evolution) == 1
            assert evolution[0]["entity"] == "dataset"
            assert evolution[0]["entity_id"] == "ds_2"


class TestCamelCaseFKEdgeCases:
    """Edge case tests for camelCase FK detection."""

    def test_get_parent_info_camelcase_fk_with_non_str_int_value(self):
        """Should return None for parent_id when camelCase FK value is not str/int."""
        row = {"id": 1, "name": "test", "companyId": None}
        parent_entity, parent_id = _get_parent_info(row, "user", None)
        assert parent_entity is None
        assert parent_id is None

        row2 = {"id": 1, "name": "test", "companyId": ["list"]}
        parent_entity2, parent_id2 = _get_parent_info(row2, "user", None)
        assert parent_entity2 is None
        assert parent_id2 is None
